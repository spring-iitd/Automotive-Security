import os
import math
import openpyxl

HEADER = ["round", "model", "I", "M", "Pi", "Pm", "D", "D_left",
          "TP", "FP", "TN", "FN", "total_attack", "total_benign",
          "ASR", "Evasion_Rate"]

# Column indices (0-based) for lookup
_COL_ROUND = 0
_COL_MODEL = 1
_COL_TP    = 8
_COL_FN    = 11


def _budgets(round_num, n):
    """Return (I, M, Pi, Pm, D) for this round. n = total attack frames."""
    n4 = math.ceil(0.25 * n) if n > 0 else 0
    table = {
        -1: (0,  0,   0,   0,  0),
        0:  (0,  0,   0,   0,  1),
        1:  (n,  0,   5,   0,  0),
        2:  (0,  n4,  5,   0,  0),
        3:  (0,  0,   n,   n4, 0),
    }
    return table.get(round_num, (0, 0, 0, 0, 0))


def _read_baseline(ws, model_name):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_COL_ROUND] == -1 and row[_COL_MODEL] == model_name:
            return int(row[_COL_TP]), int(row[_COL_FN])
    raise ValueError(
        f"Baseline row not found for model '{model_name}'. "
        "Run round=-1 evaluation first."
    )


def _find_existing_row(ws, round_num, model_name):
    """Return 1-based row index if (round, model) exists, else None."""
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[_COL_ROUND] == round_num and row[_COL_MODEL] == model_name:
            return i
    return None


def log_result(excel_path, round_num, model_name, TP, TN, FP, FN, I=None, M=None, Pi=None, Pm=None, D=None, D_left=None):
    """
    Append or update one result row in the shared Excel file.

    Parameters
    ----------
    excel_path  : str   path to spoof_CH_results.xlsx (created if absent)
    round_num   : int   -1 = baseline, 0-3 = adversarial rounds
    model_name  : str   e.g. "RI_ResNet", "ResNet50"
    TP, TN, FP, FN : int  confusion matrix values (frame-level)
    I, M, Pi, Pm, D : int  actual packet counts from attack script (optional)
    D_left : int  attack packets surviving deletion (optional, defaults to TP+FN)
    """
    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADER)

    total_attack = TP + FN
    total_benign = TN + FP

    # Use actual values from attack script if provided, else compute from formula
    if I is None or M is None or Pi is None or Pm is None or D is None:
        I_bud, M_bud, Pi_bud, Pm_bud, D_bud = _budgets(round_num, total_attack)
        I = I if I is not None else I_bud
        M = M if M is not None else M_bud
        Pi = Pi if Pi is not None else Pi_bud
        Pm = Pm if Pm is not None else Pm_bud
        D = D if D is not None else D_bud

    # D_left = attack packets remaining after deletion = total_attack - D
    if D_left is None:
        D_left = total_attack - D if round_num >= 0 else total_attack

    if round_num == -1:
        ASR = 0.0
    else:
        baseline_TP, baseline_FN = _read_baseline(ws, model_name)
        ASR = (baseline_TP - TP) / baseline_TP if baseline_TP > 0 else 0.0

    evasion = FN / (FN + TP) if (FN + TP) > 0 else 0.0

    new_row = [round_num, model_name, I, M, Pi, Pm, D, D_left,
               TP, FP, TN, FN, total_attack, total_benign,
               round(ASR, 6), round(evasion, 6)]

    existing = _find_existing_row(ws, round_num, model_name)
    if existing is not None:
        for col_idx, val in enumerate(new_row, start=1):
            ws.cell(row=existing, column=col_idx, value=val)
    else:
        ws.append(new_row)

    wb.save(excel_path)
    print(f"[result_logger] Saved round={round_num} model={model_name} "
          f"ASR={ASR:.4f} Evasion={evasion:.4f} -> {excel_path}")
